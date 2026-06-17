#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Генератор HTML-дашборда маркетинга HEYS (план 22, п.5.1).

Читает: 00_Сводная_панель.xlsx (Сводка, Конкуренты, Telegram, KPI-трекер),
22_План (единый источник статусов задач), 25_Roadmap (гейты), 29_Аудит
(сводные таблицы конкурентов), 30_Конкурентные_решения (конкурентные решения;
статусы пунктов с задачей «22 N.N» подтягиваются из 22), 24_посты
(контент-батч).
Пишет: 00_Дашборд.html (самодостаточный, офлайн; вкладки:
Обзор / Конкуренты / Telegram / План 22).

При сломанной структуре источников падает с ненулевым кодом — pre-commit
хук тогда блокирует коммит рассинхронизированного дашборда.

Запуск:  python3 маркетинг/tools/build_dashboard.py
"""
import datetime
import html
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent  # маркетинг/
README_TEXT = (ROOT / 'README.md').read_text(encoding='utf-8')


def esc(v):
    return html.escape(str(v)) if v is not None else ''


def parse_md_table(lines, start_idx):
    """Читает markdown-таблицу начиная с строки заголовка."""
    head = [c.strip().replace('**', '') for c in
            lines[start_idx].strip().strip('|').split('|')]
    rows = []
    for line in lines[start_idx + 1:]:
        if not line.strip().startswith('|'):
            break
        cells = [c.strip().replace('**', '') for c in
                 line.strip().strip('|').split('|')]
        if set(cells[0]) <= {'-', ' ', ':'}:
            continue
        rows.append(cells)
    return head, rows


def clean_md_cell(v):
    return re.sub(r'\*\*', '', str(v or '').strip())


def parse_bullets(block, title):
    m = re.search(rf'{re.escape(title)}:\n\n(.*?)(?=\n\n[A-ЯA-Z][^:\n]+:|\n### |\n---|\Z)',
                  block, re.S)
    if not m:
        return []
    items = []
    for ln in m.group(1).splitlines():
        if ln.startswith('- '):
            items.append(ln[2:].strip())
        elif items and ln.startswith('  '):
            items[-1] += ' ' + ln.strip()
    return items


def parse_line_field(block, name):
    lines = block.splitlines()
    for idx, ln in enumerate(lines):
        prefix = f'{name}:'
        if not ln.startswith(prefix):
            continue
        parts = [ln[len(prefix):].strip()]
        for next_ln in lines[idx + 1:]:
            if not next_ln.strip():
                break
            if re.match(r'^[^:\n]{2,60}:\s+', next_ln):
                break
            parts.append(next_ln.strip())
        return ' '.join(p for p in parts if p).strip()
    return ''


def parse_table_after_heading(text, heading):
    lines = text.splitlines()
    start_idx = next((i for i, ln in enumerate(lines) if ln.strip() == heading), None)
    if start_idx is None:
        return [], []
    table_idx = next((i for i in range(start_idx + 1, len(lines))
                      if lines[i].strip().startswith('|')), None)
    if table_idx is None:
        return [], []
    return parse_md_table(lines, table_idx)


def parse_table_after_heading_prefix(text, heading_prefix):
    lines = text.splitlines()
    start_idx = next((i for i, ln in enumerate(lines)
                      if ln.strip().startswith(heading_prefix)), None)
    if start_idx is None:
        return [], []
    table_idx = next((i for i in range(start_idx + 1, len(lines))
                      if lines[i].strip().startswith('|')), None)
    if table_idx is None:
        return [], []
    return parse_md_table(lines, table_idx)


def parse_paragraph_after_heading(text, heading):
    lines = text.splitlines()
    start_idx = next((i for i, ln in enumerate(lines) if ln.strip() == heading), None)
    if start_idx is None:
        return ''
    parts = []
    for ln in lines[start_idx + 1:]:
        if ln.startswith('#') or ln == '---':
            break
        if ln.strip():
            parts.append(ln.strip())
    return ' '.join(parts)


wb = openpyxl.load_workbook(ROOT / '00_Сводная_панель.xlsx', data_only=True)

# ---------- Сводка ----------
sv = wb['Сводка']
tariffs = [[sv.cell(row=r, column=c).value for c in range(2, 7)]
           for r in range(12, 16) if sv.cell(row=r, column=2).value]
tariff_note = sv['B17'].value or ''
eco = {'base': sv['B21'].value, 'real': sv['D21'].value, 'goal': sv['F21'].value,
       'note': sv['B22'].value or ''}
sut, pos = sv['B5'].value or '', sv['B8'].value or ''

kpi = []
kws = wb['KPI-трекер']
for r in range(4, kws.max_row + 1):
    if kws.cell(row=r, column=1).value:
        kpi.append({'name': kws.cell(row=r, column=1).value,
                    'goal': kws.cell(row=r, column=2).value,
                    'dir': kws.cell(row=r, column=3).value,
                    'fact': kws.cell(row=r, column=4).value})

# ---------- Конкуренты ----------
cw = wb['Конкуренты']
calzen_short = cw['A5'].value or ''
comp_head = [cw.cell(row=9, column=c).value for c in range(1, 7)]
comp_rows = [[cw.cell(row=r, column=c).value for c in range(1, 7)]
             for r in range(10, 20) if cw.cell(row=r, column=1).value]
land_rows = [[cw.cell(row=r, column=1).value, cw.cell(row=r, column=2).value,
              cw.cell(row=r, column=4).value, cw.cell(row=r, column=5).value]
             for r in range(23, 31) if cw.cell(row=r, column=1).value]
top5 = [cw.cell(row=r, column=1).value for r in range(33, 42)
        if cw.cell(row=r, column=1).value]
no_copy = cw['A44'].value or ''
# Глубокий аудит 2026-06-10 (строки 46+)
audit_head = cw['A47'].value or ''
human_rows = [[cw.cell(row=r, column=c).value for c in range(1, 6)]
              for r in range(51, 57) if cw.cell(row=r, column=1).value]
price_ladder = cw['A59'].value or ''
landing_patterns = [cw.cell(row=r, column=1).value for r in range(62, 67)
                    if cw.cell(row=r, column=1).value]

# ---------- контент-батч из 24 ----------
posts, post_calendar, post_prereq = [], '', ''
posts_path = ROOT / '24_Telegram_посты_батч1.md'
if posts_path.exists():
    p_text = posts_path.read_text(encoding='utf-8')
    posts = re.findall(r'^## (П\d+) — ([^\n*]+)', p_text, re.M)
    p_lines = p_text.splitlines()
    cal_idx = next((i for i, ln in enumerate(p_lines)
                    if ln.startswith('|') and 'Нед' in ln), None)
    if cal_idx is not None:
        _, cal_rows = parse_md_table(p_lines, cal_idx)
        post_calendar = ('<tr><th>Нед</th><th>Пн</th><th>Ср</th><th>Пт</th>'
                         '<th>Вс</th></tr>' +
                         ''.join('<tr>' + ''.join(f'<td>{esc(c)}</td>' for c in r)
                                 + '</tr>' for r in cal_rows))
    m = re.search(r'Prerequisite публикации П0:(.*?)(?=\n\n|\Z)', p_text, re.S)
    if m:
        post_prereq = ' '.join(m.group(1).replace('>', ' ').split())

# ---------- Telegram ----------
tg = wb['Telegram']
tg_role = tg['B5'].value or ''
tg_rubrics = [[tg.cell(row=r, column=2).value, tg.cell(row=r, column=3).value,
               tg.cell(row=r, column=5).value]
              for r in range(9, 15) if tg.cell(row=r, column=2).value]
tg_blocks = [('Ритм ведения', tg['B17'].value), ('Воронка', tg['B25'].value),
             ('Метрики', tg['B28'].value), ('Легал (ERID)', tg['B31'].value),
             ('Чего не делать', tg['B34'].value)]
tg_promo = [tg['B20'].value, tg['B21'].value, tg['B22'].value]
tg_start14_rows = []
tg_ad_examples = []
tg_channel_examples = []
tg_contour_sources = ([], [])
tg_contour_layers = ([], [])
tg_contour_cycle = ([], [])
tg_contour_formula = ''
tg_playbook_path = ROOT / '14_Telegram_плейбук.md'
if tg_playbook_path.exists():
    tg_playbook_text = tg_playbook_path.read_text(encoding='utf-8')
    tg_playbook_lines = tg_playbook_text.splitlines()
    start14_idx = next((i for i, ln in enumerate(tg_playbook_lines)
                        if ln.startswith('|') and 'День' in ln and 'Действие' in ln), None)
    if start14_idx is not None:
        _, tg_start14_rows = parse_md_table(tg_playbook_lines, start14_idx)
    for m in re.finditer(r'^### (TG-AD-\d+) — ([^\n]+)\n(.*?)(?=^### |\n---\n\n## |\Z)',
                         tg_playbook_text, re.M | re.S):
        body = m.group(3)
        tg_ad_examples.append({
            'id': m.group(1),
            'title': m.group(2).strip(),
            'screenshot': parse_line_field(body, 'Скрин'),
            'format': parse_line_field(body, 'Формат'),
            'offer': parse_line_field(body, 'Оффер'),
            'pattern': parse_line_field(body, 'Паттерн'),
            'scores': parse_line_field(body, 'Оценки'),
            'conclusion': parse_line_field(body, 'Вывод для HEYS'),
            'take': parse_bullets(body, 'Что взять для HEYS'),
            'avoid': parse_bullets(body, 'Чего избегать'),
        })
    for m in re.finditer(r'^### (TG-CH-\d+) — ([^\n]+)\n(.*?)(?=^### |\n---\n\n## |\Z)',
                         tg_playbook_text, re.M | re.S):
        body = m.group(3)
        tg_channel_examples.append({
            'id': m.group(1),
            'title': m.group(2).strip(),
            'link': parse_line_field(body, 'Ссылка'),
            'niche': parse_line_field(body, 'Ниша'),
            'audience': parse_line_field(body, 'Аудитория'),
            'offer': parse_line_field(body, 'Оффер'),
            'pattern': parse_line_field(body, 'Паттерн'),
            'scores': parse_line_field(body, 'Оценки'),
            'conclusion': parse_line_field(body, 'Вывод для HEYS'),
            'take': parse_bullets(body, 'Что взять для HEYS'),
            'avoid': parse_bullets(body, 'Чего избегать'),
        })
    tg_contour_sources = parse_table_after_heading(tg_playbook_text, '### Что реально берём из каналов')
    tg_contour_layers = parse_table_after_heading(tg_playbook_text, '### Сборка в один контур')
    tg_contour_cycle = parse_table_after_heading(tg_playbook_text, '### Канальный цикл на 4 недели')
    tg_contour_formula = parse_paragraph_after_heading(tg_playbook_text, '### Итоговая формула')

# ---------- сводные таблицы из 29 (приложения / лендинги) ----------
audit_apps, audit_landings = ([], []), ([], [])
adopt_landing, adopt_product = ([], []), ([], [])
no_copy_decided = ''
audit_path = ROOT / '29_Аудит_конкурентов_глубокий.md'
if audit_path.exists():
    a_text = audit_path.read_text(encoding='utf-8')
    a_lines = a_text.splitlines()
    in_landing_sec = in_product_sec = False
    for i, ln in enumerate(a_lines):
        if ln.startswith('### В ЛЕНДИНГ'):
            in_landing_sec, in_product_sec = True, False
        elif ln.startswith('### В ПРОДУКТ'):
            in_landing_sec, in_product_sec = False, True
        elif ln.startswith('#'):
            in_landing_sec = in_product_sec = False
        if ln.startswith('|') and 'Ввод еды' in ln:
            audit_apps = parse_md_table(a_lines, i)
        elif ln.startswith('|') and 'Hero-фрейм' in ln:
            audit_landings = parse_md_table(a_lines, i)
        elif ln.startswith('|') and 'У кого подсмотрено' in ln:
            if in_landing_sec:
                adopt_landing = parse_md_table(a_lines, i)
            elif in_product_sec:
                adopt_product = parse_md_table(a_lines, i)
    m = re.search(r'### НЕ КОПИРУЕМ[^\n]*\n+(.*?)\n\n---', a_text, re.S)
    if m:
        no_copy_decided = ' '.join(m.group(1).split())

# ---------- конкурентные решения из 30 ----------
imap_landing, imap_product = ([], []), ([], [])
imap_rule = imap_no_copy = ''
imap_path = ROOT / '30_Конкурентные_решения.md'
if imap_path.exists():
    i_text = imap_path.read_text(encoding='utf-8')
    i_lines = i_text.splitlines()
    cur = None
    for i, ln in enumerate(i_lines):
        if (ln.startswith('## В ЛЕНДИНГ')
                or ln.startswith('## Что заимствуем')
                or ln.startswith('## Что переносим')):
            cur = 'L'
        elif ln.startswith('## В ПРИЛОЖЕНИЕ') or ln.startswith('## После релиза'):
            cur = 'P'
        elif ln.startswith('## '):
            cur = None
        if ln.startswith('|') and 'Почему' in ln:
            if cur == 'L':
                imap_landing = parse_md_table(i_lines, i)
            elif cur == 'P':
                imap_product = parse_md_table(i_lines, i)
    m = re.search(r'\*\*Правило актуализации:\*\*(.*?)\n\n', i_text, re.S)
    if m:
        imap_rule = ' '.join(m.group(1).replace('>', ' ').split())
    m = re.search(r'## Не копируем[^\n]*\n\n(.*?)\n\n---', i_text, re.S)
    if m:
        imap_no_copy = ' '.join(m.group(1).split())


# ---------- статусы этапов из 22 ----------
TASK_ID_RE = re.compile(r'^(\d+[A-ZА-Я]?\.\d+(?:\.\d+)?)')


def parse_task_id(cell):
    m = TASK_ID_RE.match(cell.strip())
    return m.group(1) if m else None


def is_top_level_task(tid):
    return re.fullmatch(r'\d+[A-ZА-Я]?\.\d+', tid) is not None


plan_text = (ROOT / '22_План_реализации_маркетинга.md').read_text(encoding='utf-8')
release_steps = parse_table_after_heading_prefix(plan_text, '## Релизные ступени')
stages = []
task_status = {}  # '3.7' -> '✅'/'🟡'/'⬜' — единый источник статусов задач
for m in re.finditer(r'^## (Этап \d[^\n]*)\n(.*?)(?=^## |\Z)', plan_text, re.M | re.S):
    tasks = []
    task_by_id = {}
    current_status_idx = None
    current_owner_idx = None
    for line in m.group(2).splitlines():
        if not line.strip().startswith('|'):
            current_status_idx = None
            current_owner_idx = None
            continue
        cells = [clean_md_cell(c) for c in line.strip().strip('|').split('|')]
        if len(cells) < 4 or set(cells[0]) <= {'-', ' ', ':'}:
            continue
        if 'Статус' in cells:
            current_status_idx = cells.index('Статус')
            current_owner_idx = cells.index('Codex/помощь') if 'Codex/помощь' in cells else None
            continue
        tid = parse_task_id(cells[0])
        if not tid:
            continue
        if current_status_idx is None or current_status_idx >= len(cells):
            continue
        st = cells[current_status_idx]
        if '✅' not in st and '🟡' not in st and '⬜' not in st:
            continue
        status = '✅' if '✅' in st else ('🟡' if '🟡' in st else '⬜')
        owner = cells[current_owner_idx] if current_owner_idx is not None and current_owner_idx < len(cells) else ''
        item = {'id': tid, 'name': cells[1], 'status': status, 'owner': owner, 'subtasks': []}
        if is_top_level_task(tid):
            task_status[tid] = status
            tasks.append(item)
            task_by_id[tid] = item
        else:
            parent_id = '.'.join(tid.split('.')[:2])
            if parent_id in task_by_id:
                task_by_id[parent_id]['subtasks'].append(item)
    if tasks:
        t = m.group(1)
        stages.append({'title': t.split('—')[0].strip(),
                       'desc': t.split('—')[1].strip() if '—' in t else '',
                       'done': sum(1 for r in tasks if r['status'] == '✅')
                       + 0.5 * sum(1 for r in tasks if r['status'] == '🟡'),
                       'total': len(tasks),
                       'tasks': tasks})

# ---------- гейты из 25 ----------
gates = []
roadmap_text = (ROOT / '25_Roadmap_Ф0_Ф1.md').read_text(encoding='utf-8')
release_subgates = parse_table_after_heading_prefix(roadmap_text, '## Подраздел ПДн/RKN release gate')
lines = roadmap_text.splitlines()
start = next((i for i, ln in enumerate(lines) if 'Гейт / метрика' in ln), None)
if start is not None:
    for line in lines[start + 1:]:
        if not line.strip().startswith('|'):
            break
        cells = [c.strip() for c in line.strip().strip('|').split('|')]
        if len(cells) >= 7 and not set(cells[0]) <= {'-', ' '}:
            gates.append({'name': re.sub(r'\*\*', '', cells[0]), 'goal': cells[1],
                          'trigger': cells[2], 'action': cells[3],
                          'status': cells[6]})


def chip(s):
    if '✅' in s:
        return '<span class="chip ok">готово</span>'
    if '🟡' in s:
        return '<span class="chip mid">в работе</span>'
    return '<span class="chip wait">ожидает</span>'


now = datetime.datetime.now()
today = now.strftime('%Y-%m-%d %H:%M')
gen_epoch_ms = int(now.timestamp() * 1000)
total_done = sum(s['done'] for s in stages)
total_all = sum(s['total'] for s in stages)
plan_pct = round(total_done / total_all * 100) if total_all else 0

# ---------- фрагменты: Обзор ----------
tariff_cards = ''
for t in tariffs:
    feat = ' featured' if t[0] == 'Pro' else (' archive' if 'Lite' in str(t[0]) else '')
    tariff_cards += (f'<div class="tariff{feat}"><div class="t-name">{esc(t[0])}</div>'
                     f'<div class="t-price">{esc(t[1])}</div>'
                     f'<div class="t-row"><span>Куратор</span><b>{esc(t[2])}</b></div>'
                     f'<div class="t-row"><span>AI</span><b>{esc(t[3])}</b></div>'
                     f'<div class="t-row"><span>Удержание</span><b>{esc(t[4])}</b></div></div>')

gate_rows = ''.join(
    f'<tr><td>{esc(g["name"])}</td><td class="num">{esc(g["goal"])}</td>'
    f'<td class="dim">{esc(g["trigger"])}</td><td class="dim">{esc(g["action"])}</td>'
    f'<td>{chip(g["status"])}</td></tr>' for g in gates)


def md_inline(text):
    text = html.unescape(str(text))
    out = esc(text)
    out = re.sub(r'`([^`]+)`', r'<code>\1</code>', out)
    out = re.sub(r'\*\*([^*]+)\*\*', r'<strong>\1</strong>', out)
    out = re.sub(r'\[([^\]]+)\]\(([^)]+)\)', r'<a href="\2">\1</a>', out)
    return out


def task_tone_from_cells(cells, status_idx=None, owner_idx=None):
    st = cells[status_idx] if status_idx is not None and status_idx < len(cells) else ''
    if '✅' in st:
        return 'task-done'
    owner = cells[owner_idx].lower() if owner_idx is not None and owner_idx < len(cells) else ''
    needs_user = any(x in owner for x in ('тво', 'тоб', 'оба', 'codex + ты', 'ты'))
    if needs_user:
        return 'task-user'
    if 'codex сам' in owner or owner in ('security', 'codex'):
        return 'task-codex'
    return 'task-user' if ('🟡' in st or '⬜' in st) else ''


def md_table_separator(cells):
    return cells and all(re.fullmatch(r':?-{2,}:?', c.strip()) for c in cells if c.strip())


def render_plan_table(table_lines):
    rows = [[c.strip() for c in line.strip().strip('|').split('|')]
            for line in table_lines]
    if not rows:
        return ''
    head = rows[0]
    status_idx = head.index('Статус') if 'Статус' in head else None
    owner_idx = head.index('Codex/помощь') if 'Codex/помощь' in head else None
    body = rows[2:] if len(rows) > 1 and md_table_separator(rows[1]) else rows[1:]
    out = ['<div class="plan-md-table-wrap"><table class="plan-md-table"><thead><tr>']
    out.extend(f'<th>{md_inline(c)}</th>' for c in head)
    out.append('</tr></thead><tbody>')
    for row in body:
        cls = task_tone_from_cells(row, status_idx, owner_idx)
        out.append(f'<tr class="{cls}">')
        out.extend(f'<td>{md_inline(c)}</td>' for c in row)
        out.append('</tr>')
    out.append('</tbody></table></div>')
    return ''.join(out)


def render_plan_markdown(text):
    lines = text.splitlines()
    out, para = [], []
    list_type = None

    def flush_para():
        nonlocal para
        if para:
            out.append('<p>' + md_inline(' '.join(para)) + '</p>')
            para = []

    def close_list():
        nonlocal list_type
        if list_type:
            out.append(f'</{list_type}>')
            list_type = None

    i = 0
    while i < len(lines):
        line = lines[i]
        stripped = line.strip()

        if not stripped:
            flush_para()
            close_list()
            i += 1
            continue

        if stripped.startswith('|'):
            flush_para()
            close_list()
            table_lines = []
            while i < len(lines) and lines[i].strip().startswith('|'):
                table_lines.append(lines[i])
                i += 1
            out.append(render_plan_table(table_lines))
            continue

        h = re.match(r'^(#{1,6})\s+(.+)$', stripped)
        if h:
            flush_para()
            close_list()
            level = len(h.group(1))
            out.append(f'<h{level} class="plan-md-h{level}">{md_inline(h.group(2))}</h{level}>')
            i += 1
            continue

        if stripped == '---':
            flush_para()
            close_list()
            out.append('<hr class="plan-md-hr">')
            i += 1
            continue

        if stripped.startswith('>'):
            flush_para()
            close_list()
            quote = []
            while i < len(lines) and lines[i].strip().startswith('>'):
                quote.append(lines[i].strip().lstrip('>').strip())
                i += 1
            out.append('<blockquote>' + ''.join(
                '<p>' + md_inline(q) + '</p>' for q in quote if q
            ) + '</blockquote>')
            continue

        li = re.match(r'^[-*]\s+(.+)$', stripped)
        oli = re.match(r'^\d+[.]\s+(.+)$', stripped)
        if li or oli:
            flush_para()
            wanted = 'ul' if li else 'ol'
            if list_type != wanted:
                close_list()
                out.append(f'<{wanted}>')
                list_type = wanted
            out.append('<li>' + md_inline((li or oli).group(1)) + '</li>')
            i += 1
            continue

        if list_type and line.startswith(' ') and out and out[-1].endswith('</li>'):
            out[-1] = out[-1][:-5] + ' ' + md_inline(stripped) + '</li>'
            i += 1
            continue

        para.append(stripped)
        i += 1

    flush_para()
    close_list()
    return ''.join(out)


plan_full_html = render_plan_markdown(plan_text)

release_steps_rows = ''
release_step_by_name = {}
if release_steps[1]:
    for r in release_steps[1]:
        if len(r) < 5:
            continue
        step_name = clean_md_cell(r[0])
        release_step_by_name[step_name] = r
        release_steps_rows += (
            f'<tr><td>{md_inline(r[0])}</td>'
            f'<td>{md_inline(r[1])}</td>'
            f'<td class="dim">{md_inline(r[2])}</td>'
            f'<td>{md_inline(r[4])}</td></tr>')

s1_blocker_items = ''
s1_row = next((r for name, r in release_step_by_name.items()
               if name.startswith('S1:')), None)
if s1_row and len(s1_row) >= 3:
    blockers = [b.strip() for b in re.split(r';\s*', clean_md_cell(s1_row[2])) if b.strip()]
    s1_blocker_items = ''.join(f'<li>{md_inline(b)}</li>' for b in blockers)

release_subgate_rows = ''
if release_subgates[1]:
    for r in release_subgates[1]:
        if len(r) < 4:
            continue
        release_subgate_rows += (
            f'<tr><td>{md_inline(r[0])}</td><td>{md_inline(r[1])}</td>'
            f'<td>{md_inline(r[2])}</td><td class="dim">{md_inline(r[3])}</td></tr>')

kpi_rows = ''
for k in kpi:
    goal = k['goal']
    if isinstance(goal, float) and goal <= 1:
        goal = f'{goal:.0%}'
    fact = k['fact'] if k['fact'] not in (None, '') else '—'
    kpi_rows += (f'<div class="kpi{" k-wait" if fact == "—" else ""}'
                 f'"><div class="k-name">{esc(k["name"])}</div>'
                 f'<div class="k-val">{esc(fact)}</div>'
                 f'<div class="k-goal">цель {esc(k["dir"] or "")} {esc(goal)}</div></div>')

# ---------- фрагменты: Конкуренты ----------
comp_table = '<tr>' + ''.join(f'<th>{esc(h)}</th>' for h in comp_head) + '</tr>'
for r in comp_rows:
    comp_table += ('<tr><td class="dim">' + esc(r[0]) + '</td><td class="hl">' +
                   esc(r[1]) + '</td>' +
                   ''.join(f'<td>{esc(x)}</td>' for x in r[2:]) + '</tr>')

land_table = ('<tr><th>Блок</th><th>CalZen</th><th>Наш</th><th>Что перенять</th></tr>' +
              ''.join('<tr><td class="dim">' + esc(r[0]) + '</td><td>' + esc(r[1]) +
                      '</td><td>' + esc(r[2]) + '</td><td class="hl">' + esc(r[3]) +
                      '</td></tr>' for r in land_rows))

top5_rows = ''.join(f'<div class="prio top5"><div class="p-num acc2">{i}</div>'
                    f'<div>{esc(re.sub(r"^[0-9]+[.][ ]*", "", str(p)))}</div></div>'
                    for i, p in enumerate(top5, 1))

# ---------- фрагменты: Telegram ----------
human_table = ''.join(
    '<tr>' + ''.join(
        ('<td class="hl">' if i == 0 and 'HEYS' in str(r[0]) else '<td>') + esc(x) + '</td>'
        for i, x in enumerate(r)) + '</tr>'
    for r in human_rows)
landing_pat_rows = ''.join(
    f'<div class="prio top5"><div class="p-num acc2">{i}</div>'
    f'<div>{esc(re.sub(r"^[0-9]+[.][ ]*", "", str(p)))}</div></div>'
    for i, p in enumerate(landing_patterns, 1))


def render_audit_table(head_rows):
    head, rows = head_rows
    if not rows:
        return '<p class="dim">таблица не найдена в 29</p>'
    out = '<tr>' + ''.join(f'<th>{esc(h)}</th>' for h in head) + '</tr>'
    for r in rows:
        hl = ' class="hl-row"' if 'HEYS' in r[0] else ''
        out += f'<tr{hl}>' + ''.join(f'<td>{esc(c)}</td>' for c in r) + '</tr>'
    return out


audit_apps_table = render_audit_table(audit_apps)
audit_landings_table = render_audit_table(audit_landings)


def render_adopt(head_rows):
    _, rows = head_rows
    out = ''
    for r in rows:
        if len(r) < 4:
            continue
        st = r[3]
        cls = 'ok' if '✅' in st else ('mid' if '🟡' in st else 'wait')
        out += (f'<div class="adopt"><span class="chip {cls}">{esc(st)}</span>'
                f'<div><b>{esc(r[0])}</b>'
                f'<div class="label">{esc(r[1])} → {esc(r[2])}</div></div></div>')
    return out or '<p class="dim">нет данных (29 §10)</p>'


adopt_landing_html = render_adopt(adopt_landing)
adopt_product_html = render_adopt(adopt_product)


def render_imap(head_rows):
    _, rows = head_rows
    out = ('<tr><th>№</th><th>Что</th><th>Почему (29)</th><th>Задача</th>'
           '<th>Очередь</th><th>Статус</th></tr>')
    done = 0
    for r in rows:
        if len(r) < 6:
            continue
        st = imap_status(r)
        cls = 'ok' if '✅' in st else ('mid' if '🟡' in st else 'wait')
        if '✅' in st:
            done += 1
        label = 'готово' if '✅' in st else ('в работе' if '🟡' in st else 'ожидает')
        out += (f'<tr><td class="num">{esc(r[0])}</td><td><b>{esc(r[1])}</b></td>'
                f'<td class="dim">{esc(r[2])}</td><td>{esc(r[3])}</td>'
                f'<td class="dim">{esc(r[4])}</td>'
                f'<td><span class="chip {cls}">{label}</span></td></tr>')
    return out, done, len(rows)


def imap_task_id(row):
    if len(row) < 4:
        return None
    tm = re.search(r'22[`»\s]*\s*(\d+\.\d+)', row[3])
    return tm.group(1) if tm else None


def imap_status(row):
    fallback = row[5] if len(row) > 5 else '⬜'
    tid = imap_task_id(row)
    return task_status.get(tid, fallback)


imap_landing_table, imap_l_done, imap_l_total = render_imap(imap_landing)
imap_product_table, imap_p_done, imap_p_total = render_imap(imap_product)


def render_plain_table(head_rows, empty_msg):
    head, rows = head_rows
    if not rows:
        return f'<p class="dim">{esc(empty_msg)}</p>'
    out = '<tr>' + ''.join(f'<th>{esc(h)}</th>' for h in head) + '</tr>'
    for r in rows:
        out += '<tr>' + ''.join(f'<td>{esc(c)}</td>' for c in r) + '</tr>'
    return '<table>' + out + '</table>'


tg_contour_sources_table = render_plain_table(tg_contour_sources, 'контур источников не найден в 14 §13')
tg_contour_layers_table = render_plain_table(tg_contour_layers, 'слои контура не найдены в 14 §13')
tg_contour_cycle_table = render_plain_table(tg_contour_cycle, 'цикл на 4 недели не найден в 14 §13')

tg_rubric_rows = ('<tr><th>Рубрика</th><th>Цель</th><th>Частота</th></tr>' +
                  ''.join(f'<tr><td>{esc(r[0])}</td><td class="dim">{esc(r[1])}</td>'
                          f'<td class="num">{esc(r[2])}</td></tr>' for r in tg_rubrics))
tg_block_cards = ''.join(f'<div class="card"><h2>{esc(t)}</h2><p class="sm">{esc(b)}</p></div>'
                         for t, b in tg_blocks if b)
tg_promo_list = ''.join(f'<li>{esc(p)}</li>' for p in tg_promo if p)
tg_start14_table = ('<tr><th>День</th><th>Действие</th></tr>' +
                    ''.join(f'<tr><td class="num">{esc(r[0])}</td><td>{esc(r[1])}</td></tr>'
                            for r in tg_start14_rows))
tg_ad_cards = ''
for ad in tg_ad_examples:
    take = '<ul>' + ''.join(f'<li>{esc(x)}</li>' for x in ad['take'][:3]) + '</ul>'
    avoid = '<ul>' + ''.join(f'<li>{esc(x)}</li>' for x in ad['avoid'][:3]) + '</ul>'
    screenshot = (f'<a class="tg-shot-link" href="{esc(ad["screenshot"])}" target="_blank" '
                  f'rel="noopener" title="Открыть скрин в полном размере">'
                  f'<img class="tg-shot" src="{esc(ad["screenshot"])}" alt="{esc(ad["id"])}"></a>'
                  if ad['screenshot'] else '')
    conclusion = (f'<div class="tg-conclusion"><b>Вывод для HEYS:</b> '
                  f'{esc(ad["conclusion"])}</div>') if ad.get('conclusion') else ''
    tg_ad_cards += (
        f'<article class="tg-example-card tg-example-card--ad">'
        f'<div class="tg-example-media">{screenshot}'
        f'<div class="tg-example-id">{esc(ad["id"])}</div>'
        f'<div class="dim">{esc(ad["title"])}</div></div>'
        f'<div class="tg-example-main"><div class="tg-example-head">'
        f'<div><h3>{esc(ad["format"])}</h3><p>{esc(ad["offer"])}</p>'
        f'<p class="label">{esc(ad["pattern"])}</p></div>'
        f'<div class="tg-score">{esc(ad["scores"])}</div></div>'
        f'{conclusion}'
        f'<div class="tg-example-cols"><div><h4>Взять для HEYS</h4>{take}</div>'
        f'<div><h4>Избегать</h4>{avoid}</div></div></div></article>')
tg_channel_cards = ''
for ch in tg_channel_examples:
    take = '<ul>' + ''.join(f'<li>{esc(x)}</li>' for x in ch['take'][:3]) + '</ul>'
    avoid = '<ul>' + ''.join(f'<li>{esc(x)}</li>' for x in ch['avoid'][:3]) + '</ul>'
    link = (f'<a href="{esc(ch["link"])}" target="_blank" rel="noopener">'
            f'{esc(ch["link"])}</a>') if ch['link'] else ''
    conclusion = (f'<div class="tg-conclusion"><b>Вывод для HEYS:</b> '
                  f'{esc(ch["conclusion"])}</div>') if ch.get('conclusion') else ''
    tg_channel_cards += (
        f'<article class="tg-example-card">'
        f'<div class="tg-example-media"><div class="tg-example-id">{esc(ch["id"])}</div>'
        f'<h3>{esc(ch["title"])}</h3><p>{link}</p></div>'
        f'<div class="tg-example-main"><div class="tg-example-head">'
        f'<div><h3>{esc(ch["niche"])}</h3><p>{esc(ch["audience"])}</p>'
        f'<p class="label">{esc(ch["offer"])}</p><p class="label">{esc(ch["pattern"])}</p></div>'
        f'<div class="tg-score">{esc(ch["scores"])}</div></div>'
        f'{conclusion}'
        f'<div class="tg-example-cols"><div><h4>Взять для HEYS</h4>{take}</div>'
        f'<div><h4>Избегать</h4>{avoid}</div></div></div></article>')

html_out = f'''<!DOCTYPE html>
<html lang="ru"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>HEYS · Панель управления маркетингом</title>
<style>
:root {{ --bg:#0b1020; --card:#121931; --card2:#0f1530; --line:#1f2a4d;
  --txt:#e8ecf8; --dim:#8b96b8; --acc:#4f8cff; --ok:#2dd4a7; --warn:#f5b14c;
  --red:#f0647c; }}
* {{ box-sizing:border-box; margin:0; }}
body {{ background:radial-gradient(1100px 500px at 80% -10%,#16224a 0%,var(--bg) 55%);
  color:var(--txt); font:13.5px/1.45 -apple-system,'SF Pro Text',Segoe UI,Roboto,sans-serif;
  padding:18px clamp(12px,3vw,36px) 40px; }}
h1 {{ font-size:clamp(17px,2.2vw,23px); letter-spacing:-.4px; }}
h2 {{ font-size:11px; text-transform:uppercase; letter-spacing:.12em;
  color:var(--dim); margin:0 0 8px; }}
.header {{ display:flex; flex-wrap:wrap; gap:10px 16px; align-items:center;
  justify-content:space-between; margin-bottom:8px; }}
.badge {{ background:linear-gradient(135deg,#4f8cff33,#2dd4a733);
  border:1px solid #4f8cff55; padding:4px 12px; border-radius:999px; font-size:12px; }}
.sub {{ color:var(--dim); font-size:12.5px; max-width:1000px; margin-bottom:12px; }}
.tabs {{ display:flex; gap:6px; margin:4px 0 16px; border-bottom:1px solid var(--line); }}
.tab {{ background:none; border:none; color:var(--dim); font:600 13px inherit;
  padding:8px 14px; cursor:pointer; border-bottom:2px solid transparent; }}
.tab.active {{ color:var(--txt); border-bottom-color:var(--acc); }}
.pane {{ display:none; }} .pane.active {{ display:block; }}
.grid {{ display:grid; gap:12px; }}
.cards3 {{ grid-template-columns:repeat(auto-fit,minmax(190px,1fr)); }}
.card {{ background:linear-gradient(180deg,var(--card),var(--card2));
  border:1px solid var(--line); border-radius:12px; padding:12px 14px; }}
.big {{ font-size:clamp(24px,3vw,34px); font-weight:700; letter-spacing:-.5px; }}
.big.ok {{ color:var(--ok); }} .big.acc {{ color:var(--acc); }}
.label {{ color:var(--dim); font-size:11.5px; }}
.sm {{ font-size:12.5px; }}
section {{ margin-top:18px; }}
table {{ width:100%; border-collapse:collapse; font-size:12.5px; }}
th {{ text-align:left; color:var(--dim); font-weight:500; font-size:10.5px;
  text-transform:uppercase; letter-spacing:.07em; padding:6px 8px;
  border-bottom:1px solid var(--line); }}
td {{ padding:6px 8px; border-bottom:1px solid #16203f; vertical-align:top; }}
td.num {{ white-space:nowrap; font-weight:600; }}
td.hl {{ color:var(--ok); font-weight:600; }}
tr.hl-row td {{ color:var(--ok); font-weight:600; background:#2dd4a70d; }}
.scrollx {{ overflow-x:auto; }} .scrollx table {{ min-width:760px; }}
.dim {{ color:var(--dim); }}
.chip {{ padding:2px 9px; border-radius:999px; font-size:11px; white-space:nowrap; }}
.chip.ok {{ background:#2dd4a722; color:var(--ok); border:1px solid #2dd4a744; }}
.chip.mid {{ background:#f5b14c22; color:var(--warn); border:1px solid #f5b14c44; }}
.chip.wait {{ background:#8b96b81e; color:var(--dim); border:1px solid #8b96b833; }}
.tariffs {{ grid-template-columns:repeat(auto-fit,minmax(160px,1fr)); }}
.tariff {{ background:var(--card); border:1px solid var(--line); border-radius:12px;
  padding:10px 12px; }}
.tariff.featured {{ border-color:var(--acc); box-shadow:0 0 0 1px var(--acc); }}
.tariff.archive {{ opacity:.5; }}
.t-name {{ font-weight:700; font-size:13px; }}
.t-price {{ font-size:19px; font-weight:700; margin-bottom:6px; color:var(--acc); }}
.tariff.featured .t-price {{ color:var(--ok); }}
.t-row {{ display:flex; justify-content:space-between; gap:8px; font-size:11.5px;
  padding:2.5px 0; border-top:1px dashed #1f2a4d; }}
.t-row span {{ color:var(--dim); }} .t-row b {{ font-weight:500; text-align:right; }}
.stage {{ display:grid; grid-template-columns:minmax(150px,1fr) 1.6fr; gap:10px;
  align-items:center; padding:6px 0; border-bottom:1px solid #16203f; font-size:12.5px; }}
.s-head {{ display:flex; flex-direction:column; }}
.s-bar {{ display:flex; align-items:center; gap:8px; }}
.bar {{ flex:1; height:6px; background:#1a2447; border-radius:99px; overflow:hidden; }}
.bar-fill {{ height:100%; background:linear-gradient(90deg,var(--acc),var(--ok)); }}
.bar-num {{ font-weight:700; font-size:12px; min-width:34px; }}
.s-count {{ font-size:11px; min-width:36px; text-align:right; }}
details.stage-d > summary {{ cursor:pointer; list-style:none; }}
details.stage-d > summary::-webkit-details-marker {{ display:none; }}
details.stage-d[open] > summary .s-head b {{ color:var(--acc); }}
.task-list {{ margin:4px 0 10px; padding-left:6px; list-style:none; }}
.task-list li {{ font-size:12px; padding:3px 0; border-bottom:1px dashed #16203f; }}
.task-list li.task-done,.subtask-list li.task-done {{ color:var(--dim); opacity:.62; }}
.task-list li.task-codex,.subtask-list li.task-codex {{ color:var(--ok); }}
.task-list li.task-user,.subtask-list li.task-user {{ color:var(--txt); }}
.task-list .chip {{ margin-right:6px; }}
.subtask-list {{ margin:4px 0 1px 22px; padding-left:0; list-style:none; }}
.subtask-list li {{ color:var(--dim); font-size:11.5px; padding:2px 0;
  border-bottom:0; }}
.subtask-list .chip {{ margin-right:6px; opacity:.92; }}
.kpis {{ grid-template-columns:repeat(auto-fit,minmax(140px,1fr)); }}
.kpi {{ background:var(--card); border:1px solid var(--line); border-radius:10px;
  padding:9px 11px; }}
.kpi.k-wait .k-val {{ color:var(--dim); }}
.k-name {{ font-size:11px; color:var(--dim); min-height:26px; }}
.k-val {{ font-size:19px; font-weight:700; }}
.k-goal {{ font-size:11px; color:var(--dim); }}
.prios {{ display:grid; gap:8px; grid-template-columns:repeat(auto-fit,minmax(330px,1fr)); }}
.prio {{ display:flex; gap:10px; background:var(--card); border:1px solid var(--line);
  border-left:3px solid var(--red); border-radius:10px; padding:9px 12px; font-size:12.5px; }}
.prio.top5 {{ border-left-color:var(--acc); }}
.adopt {{ display:flex; gap:10px; align-items:flex-start; background:var(--card);
  border:1px solid var(--line); border-radius:10px; padding:9px 12px;
  margin-bottom:8px; font-size:12.5px; }}
.adopt .chip {{ flex-shrink:0; margin-top:2px; }}
.tg-examples {{ display:grid; gap:12px; }}
.tg-example-card {{ display:grid; grid-template-columns:220px minmax(0,1fr); gap:14px;
  background:linear-gradient(180deg,var(--card),var(--card2)); border:1px solid var(--line);
  border-radius:12px; padding:12px; }}
.tg-example-media {{ min-width:0; }}
.tg-example-id {{ font-weight:800; letter-spacing:.02em; margin-bottom:4px; }}
.tg-example-media h3,.tg-example-main h3 {{ font-size:13px; margin:0 0 4px; color:var(--txt); }}
.tg-example-media p,.tg-example-main p {{ margin:0 0 5px; overflow-wrap:anywhere; }}
.tg-example-head {{ display:grid; grid-template-columns:minmax(0,1fr) 220px; gap:12px;
  align-items:start; margin-bottom:10px; }}
.tg-score {{ color:var(--ok); font-weight:700; font-size:12px; line-height:1.45;
  background:#2dd4a712; border:1px solid #2dd4a733; border-radius:8px; padding:8px; }}
.tg-conclusion {{ margin:0 0 10px; padding:9px 10px; border-radius:8px;
  background:#4f8cff14; border:1px solid #4f8cff33; font-size:12.5px; }}
.tg-contour {{ display:grid; gap:12px; }}
.tg-contour-formula {{ border-left:3px solid var(--ok); }}
.tg-contour-grid {{ display:grid; gap:12px; grid-template-columns:1.1fr .9fr; }}
.tg-contour .card {{ overflow-x:auto; }}
.tg-contour table {{ min-width:680px; }}
.tg-example-cols {{ display:grid; grid-template-columns:repeat(2,minmax(0,1fr)); gap:12px; }}
.tg-example-cols h4 {{ margin:0 0 6px; color:var(--dim); font-size:10.5px; text-transform:uppercase;
  letter-spacing:.08em; }}
.tg-example-cols ul {{ margin:0; }}
.tg-shot-link {{ display:block; width:190px; max-width:100%; margin:0 0 8px; }}
.tg-shot {{ display:block; width:100%; height:auto; border-radius:8px;
  border:1px solid #2c3a66; background:#0b1020; transition:transform .12s ease,border-color .12s ease; }}
.tg-shot-link:hover .tg-shot {{ transform:scale(1.02); border-color:var(--acc); }}
@media (max-width:760px) {{
  .stage {{ grid-template-columns:1fr; }}
  .tg-example-card,.tg-example-head,.tg-example-cols {{ grid-template-columns:1fr; }}
  .tg-contour-grid {{ grid-template-columns:1fr; }}
  .tg-shot-link {{ width:100%; max-width:260px; }}
}}
#stale-warn {{ display:none; background:#f5b14c22; border:1px solid #f5b14c66;
  color:var(--warn); border-radius:10px; padding:8px 14px; margin-bottom:14px;
  font-size:12.5px; }}
.p-num {{ font-size:16px; font-weight:800; color:var(--red); line-height:1.3; }}
.p-num.acc2 {{ color:var(--acc); }}
.cols2 {{ display:grid; gap:12px; grid-template-columns:repeat(auto-fit,minmax(300px,1fr)); }}
.plan-source {{ padding:18px 20px; overflow:auto; }}
.plan-source h1 {{ font-size:24px; margin:0 0 18px; }}
.plan-source h2 {{ font-size:20px; margin:28px 0 12px; padding-bottom:8px;
  border-bottom:1px solid var(--line); }}
.plan-source h3 {{ font-size:15px; margin:22px 0 10px; color:var(--acc); }}
.plan-source p {{ margin:0 0 10px; line-height:1.6; font-size:13px; }}
.plan-source blockquote {{ margin:12px 0; padding:10px 12px; border-left:3px solid var(--acc);
  background:#4f8cff12; border-radius:8px; color:var(--txt); }}
.plan-source blockquote p:last-child {{ margin-bottom:0; }}
.plan-source code {{ font-family:ui-monospace,SFMono-Regular,Menlo,Monaco,Consolas,"Liberation Mono",monospace;
  background:#ffffff14; border:1px solid #ffffff1c; border-radius:4px; padding:1px 4px;
  font-size:.92em; color:#dbe7ff; }}
.plan-source a {{ color:var(--acc); text-decoration:none; border-bottom:1px dashed #4f8cff88; }}
.plan-source ul,.plan-source ol {{ margin:8px 0 14px; padding-left:22px; }}
.plan-source li {{ line-height:1.55; }}
.plan-md-hr {{ border:0; border-top:1px solid var(--line); margin:18px 0; }}
.plan-md-table-wrap {{ overflow-x:auto; margin:12px 0 20px; border:1px solid var(--line);
  border-radius:10px; background:#0d1428; }}
.plan-md-table {{ min-width:860px; font-size:12px; }}
.plan-md-table th {{ background:#111a33; }}
.plan-md-table td,.plan-md-table th {{ padding:7px 9px; }}
.plan-md-table tr.task-done td {{ color:var(--dim); opacity:.62; }}
.plan-md-table tr.task-codex td {{ color:var(--ok); }}
.plan-md-table tr.task-user td {{ color:var(--txt); }}
ul {{ padding-left:18px; }} li {{ margin-bottom:4px; font-size:12.5px; }}
footer {{ margin-top:26px; color:var(--dim); font-size:11px;
  border-top:1px solid var(--line); padding-top:10px; }}
</style></head><body>

<div class="header">
  <h1>HEYS · Панель управления маркетингом</h1>
  <div class="tabs">
    <button class="tab active" data-pane="overview">Обзор</button>
    <button class="tab" data-pane="comp">Конкуренты</button>
    <button class="tab" data-pane="tg">Telegram</button>
    <button class="tab" data-pane="plan22">План 22</button>
  </div>
  <span class="badge">Фаза 0 · Pro-first · план: {plan_pct}% · собрано {today}</span>
</div>
<div id="stale-warn">⚠ Этот снимок дашборда собран больше суток назад — данные могли устареть.
Обновить: двойной клик по <b>Обновить_дашборд.command</b> (или
<code>python3 маркетинг/tools/build_dashboard.py</code>).</div>

<div class="pane active" id="overview">
<p class="sub">{esc(sut)} {esc(pos)}</p>
<div class="grid cards3">
  <div class="card"><div class="label">LTV:CAC реальный (триал = время куратора)</div>
    <div class="big ok">{esc(eco['real'])}</div><div class="label">цель {esc(eco['goal'])}</div></div>
  <div class="card"><div class="label">LTV:CAC модельный</div>
    <div class="big acc">{esc(eco['base'])}</div><div class="label">микс Self 20 / Pro 65 / Pro+ 15</div></div>
  <div class="card"><div class="label">Готовность плана (22)</div>
    <div class="big">{plan_pct}%</div><div class="label">{total_done:g} из {total_all} задач</div></div>
</div>
<p class="label" style="margin-top:6px">{esc(eco['note'])}</p>

<section><h2>Релизные ступени S0–S4</h2>
<div class="card scrollx" style="padding:4px 8px"><table>
<tr><th>Ступень</th><th>Что разрешает</th><th>Что блокирует сейчас</th><th>Статус</th></tr>
{release_steps_rows}</table></div></section>

<section class="cols2">
<div class="card"><h2>Блокеры S1 сейчас</h2><ul>{s1_blocker_items}</ul></div>
<div class="card"><h2>R0–R3 release gate</h2><div class="scrollx" style="padding:4px 0"><table>
<tr><th>Гейт</th><th>Блокирует</th><th>Статус</th><th>Следующее действие</th></tr>
{release_subgate_rows}</table></div></div>
</section>

<section><h2>Гейты Ф0 → Ф1 (roadmap 25)</h2>
<div class="card" style="padding:4px 8px"><table>
<tr><th>Гейт</th><th>Цель</th><th>🔴 Триггер</th><th>Действие при провале</th><th>Статус</th></tr>
{gate_rows}</table></div></section>

<section><h2>Экономика и тарифы (xlsx)</h2>
<div class="grid tariffs">{tariff_cards}</div>
<p class="label" style="margin-top:6px">{esc(tariff_note)}</p>
</section>

<section><h2>KPI-трекер (факт — с первого трафика)</h2>
<div class="grid kpis">{kpi_rows}</div></section>
</div>

<div class="pane" id="comp">
<p class="sub"><b>Главный вывод аудита 2026-06-10:</b> {esc(audit_head)}</p>
<p class="label">Конкурентные решения из 30 встроены сюда; живые статусы подтягиваются из плана 22.</p>
<section><h2>Карта «живой человек» (кто реально даёт)</h2>
<div class="card" style="padding:4px 8px"><table>
<tr><th>Игрок</th><th>Человек</th><th>Частота</th><th>Цена/мес</th><th>РФ-оплата</th></tr>
{human_table}
</table></div>
<p class="label" style="margin-top:6px">{esc(price_ladder)}</p></section>
<section><h2>Сводная: ПРИЛОЖЕНИЯ — все игроки (29 §4)</h2>
<div class="card scrollx" style="padding:4px 8px"><table>{audit_apps_table}</table></div></section>
<section><h2>Сводная: ЛЕНДИНГИ — все игроки (29 §5)</h2>
<div class="card scrollx" style="padding:4px 8px"><table>{audit_landings_table}</table></div></section>
<section><h2>Лендинги лидеров — 5 паттернов</h2>
{landing_pat_rows}
</section>
<p class="sub" style="margin-top:14px"><b>CalZen — главный референс:</b> {esc(calzen_short)}</p>
<section><h2>Сводная таблица сегментов</h2>
<div class="card" style="padding:4px 8px"><table>{comp_table}</table></div></section>
<section><h2>Лендинги: CalZen vs наш — что перенять</h2>
<div class="card" style="padding:4px 8px"><table>{land_table}</table></div></section>
<section><h2>Что оставить из конкурентки</h2>{top5_rows}</section>
<section><h2>Решения в плане 22 — {imap_l_done}/{imap_l_total} готово</h2>
<p class="sub">{esc(imap_rule)}</p>
<div class="card scrollx" style="padding:4px 8px"><table>{imap_landing_table}</table></div></section>
<section><h2>После релиза — {imap_p_done}/{imap_p_total} готово</h2>
<div class="card scrollx" style="padding:4px 8px"><table>{imap_product_table}</table></div></section>
<section class="card"><h2>Чего не копировать</h2><p class="sm">{esc(imap_no_copy or no_copy)}</p></section>
<p class="label" style="margin-top:10px">Детали: 12 (CalZen/лендинги) · 06 (позиционирование) · 16 (паттерны) · навигатор — 26 · решения — 30 · статус — 22.</p>
</div>

<div class="pane" id="tg">
<p class="sub">{esc(tg_role)}</p>
<section><h2>Контент-батч №1 — {len(posts)} постов готово (24)</h2>
<div class="cols2">
<div class="card" style="padding:4px 8px"><table>{post_calendar}</table>
<p class="label" style="padding:6px">⚠ {esc(post_prereq)}</p></div>
<div class="card"><ul class="task-list">
{''.join(f'<li><span class="chip wait">{esc(pid)}</span> {esc(t.strip())}</li>' for pid, t in posts)}
</ul></div></div></section>
	<div class="cols2">
	<div><h2>Рубрики (банк тем — 14, батч №1 — 24)</h2>
	<div class="card" style="padding:4px 8px"><table>{tg_rubric_rows}</table></div>
	<section><h2>Продвижение</h2><div class="card"><ul>{tg_promo_list}</ul></div></section></div>
	<div class="grid" style="align-content:start">{tg_block_cards}</div>
	</div>
	<section><h2>Первые 14 дней канала (14)</h2>
	<div class="card scrollx" style="padding:4px 8px"><table>{tg_start14_table}</table></div></section>
	<section class="tg-contour"><h2>Единый контур Telegram-канала HEYS (14 §13)</h2>
	<div class="card tg-contour-formula"><h2>Итоговая формула</h2><p class="sm">{esc(tg_contour_formula)}</p></div>
	<div class="card">{tg_contour_sources_table}</div>
	<div class="tg-contour-grid">
	<div class="card">{tg_contour_layers_table}</div>
	<div class="card">{tg_contour_cycle_table}</div>
	</div></section>
	<section><h2>Библиотека рекламных примеров (14 §11)</h2>
	<div class="tg-examples">{tg_ad_cards}</div></section>
	<section><h2>Библиотека Telegram-каналов (14 §12)</h2>
	<div class="tg-examples">{tg_channel_cards}</div></section>
	</div>

<div class="pane" id="plan22">
<p class="sub"><b>Полный источник:</b> 22_План_реализации_маркетинга.md. Цвет строк задач:
серый = закрыто, зелёный = Codex может сделать сам, белый = нужно участие основателя / доступ / внешнее решение.</p>
<section><h2>План реализации маркетинга — полный текст</h2>
<div class="card plan-source">{plan_full_html}</div></section>
</div>

<footer>Сгенерировано {today} · данные: 00_Сводная_панель.xlsx · 22_План · 25_Roadmap · 29_Аудит · 30_Решения ·
обновление: <b>Обновить_дашборд.command</b> (двойной клик) · авто на каждом коммите источников ·
<code>python3 маркетинг/tools/build_dashboard.py</code></footer>

<script>
(function () {{
  var age = Date.now() - {gen_epoch_ms};
  if (age > 24 * 3600 * 1000) document.getElementById('stale-warn').style.display = 'block';
}})();
</script>
<script>
document.querySelectorAll('.tab').forEach(function (t) {{
  t.addEventListener('click', function () {{
    document.querySelectorAll('.tab').forEach(function (x) {{ x.classList.remove('active'); }});
    document.querySelectorAll('.pane').forEach(function (x) {{ x.classList.remove('active'); }});
    t.classList.add('active');
    document.getElementById(t.dataset.pane).classList.add('active');
  }});
}});
</script>
</body></html>'''

# ---------- sanity-checks: громко падаем при сломанной структуре ----------
problems = []
numbered_md_missing_status = []
for md_path in sorted(ROOT.glob('[0-9][0-9]_*.md')):
    md_text = md_path.read_text(encoding='utf-8')
    if not re.search(r'^> \*\*Статус:\*\*', md_text, re.M):
        numbered_md_missing_status.append(md_path.name)

for cond, msg in [
    ('## Статус корпуса' in README_TEXT, 'README: нет раздела «Статус корпуса»'),
    ('Правило: статус задачи меняется только в `22`' in README_TEXT,
     'README: нет правила, что task-статусы меняются только в 22'),
    (not numbered_md_missing_status,
     'маркетинг/*.md: нет статусной шапки в ' + ', '.join(numbered_md_missing_status)),
    ('data-pane="imap"' not in html_out and 'id="imap"' not in html_out,
     'дашборд: вернулась отдельная вкладка/панель imap вместо интеграции в «Конкуренты»'),
    (len(tariffs) >= 3, 'Сводка: тарифная сетка < 3 строк (B12:F15)'),
    (eco['real'] and eco['base'], 'Сводка: пустая экономика (B21/D21)'),
    (len(release_steps[1]) >= 5, '22: релизные ступени S0–S4 не найдены или < 5 строк'),
    (len(stages) >= 5, '22_План: найдено < 5 этапов'),
    (sum(s['total'] for s in stages) >= 20, '22_План: подозрительно мало задач'),
    (len(gates) >= 7, '25_Roadmap: гейтов < 7'),
    (len(release_subgates[1]) >= 4, '25_Roadmap: release subgates R0–R3 < 4 строк'),
    (bool(s1_blocker_items), '22: S1 blockers не собраны из релизных ступеней'),
    (len(kpi) >= 10, 'KPI-трекер: метрик < 10'),
    (len(audit_apps[1]) >= 10, '29 §4: таблица приложений < 10 строк'),
    (len(audit_landings[1]) >= 10, '29 §5: таблица лендингов < 10 строк'),
    (len(human_rows) >= 4, 'xlsx Конкуренты: карта «живой человек» < 4 строк (A51+)'),
    (len(imap_landing[1]) >= 5, '30: секция заимствований < 5 пунктов'),
    (len(imap_product[1]) >= 2, '30: секция «После релиза» < 2 пунктов'),
    (len(posts) >= 5, '24: найдено < 5 постов (заголовки «## ПN — …»)'),
    (len(tg_start14_rows) >= 5, '14: план первых 14 дней Telegram < 5 строк'),
    (len(tg_ad_examples) >= 1, '14 §11: библиотека рекламных примеров пуста'),
    (len(tg_channel_examples) >= 1, '14 §12: библиотека Telegram-каналов пуста'),
    (all(ad.get('conclusion') for ad in tg_ad_examples),
     '14 §11: у каждого TG-AD-* должен быть «Вывод для HEYS»'),
    (all(ch.get('conclusion') for ch in tg_channel_examples),
     '14 §12: у каждого TG-CH-* должен быть «Вывод для HEYS»'),
    (len(tg_contour_sources[1]) >= 7, '14 §13: таблица «Что реально берём» < 7 строк'),
    (len(tg_contour_layers[1]) >= 6, '14 §13: таблица «Сборка в один контур» < 6 строк'),
    (len(tg_contour_cycle[1]) >= 4, '14 §13: цикл на 4 недели < 4 строк'),
    (bool(tg_contour_formula), '14 §13: пустая итоговая формула'),
]:
    if not cond:
        problems.append(msg)

# ВСЕ исполняемые строки конкурентных решений должны быть привязаны к живой задаче 22.
for sec_name, (_, rows) in (('Что заимствуем', imap_landing), ('После релиза', imap_product)):
    for r in rows:
        if len(r) < 6:
            problems.append(
                f'30 ({sec_name}): строка конкурентных решений должна иметь колонки '
                '№/Что/Почему/Задача/Очередь/Статус')
            continue
        tids = re.findall(r'22[`»\s]*\s*(\d+\.\d+)', r[3])
        if not tids:
            problems.append(
                f'30 ({sec_name}, {r[0]}): нет ссылки на задачу 22 N.N — '
                'статус не сможет подтянуться автоматически')
        for tid in tids:
            if tid not in task_status:
                problems.append(
                    f'30 ({sec_name}, {r[0]}): ссылка на задачу 22 {tid} — '
                    f'такой задачи нет в 22 (потеряна/переименована?)')
if problems:
    import sys
    print('❌ Структура источников сломана — дашборд НЕ пересобран:')
    for p in problems:
        print('   •', p)
    print('Почини источник (см. docstring) и перезапусти.')
    sys.exit(1)

out = ROOT / '00_Дашборд.html'
out.write_text(html_out, encoding='utf-8')
print(f'OK → {out} ({len(html_out)} bytes; этапов {len(stages)}, гейтов {len(gates)}, '
      f'KPI {len(kpi)}, конкуренты {len(comp_rows)} строк, рубрик {len(tg_rubrics)})')
